from pathlib import Path
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

# ======================================================
# CONFIGURACIÓN
# ======================================================
st.set_page_config(page_title="Registro de ventas", page_icon="🧾", layout="centered")

RUTA_ARCHIVO = Path(r"C:\Users\vamon\Documents\Aplicación registro ventas\completos_app_1.1\ventas_historicas.xlsx")
HOJA_VENTAS = "Ventas"
HOJA_PRECIOS = "Precios"

COLUMNAS_VENTAS = [
    "fecha",
    "hora",
    "nombre_comprador",
    "cantidad_promo_completo_bebida",
    "cantidad_completos_solos",
    "cantidad_bebidas_solas",
    "cantidad_cafes_solos",
    "cantidad_te_solos",
    "total_venta",
]

PRODUCTOS_ESPERADOS = [
    "promocion_completo_bebida",
    "completo_solo",
    "bebida_sola",
    "cafe_solo",
    "te_solo",   
]


# ======================================================
# FUNCIONES
# ======================================================
def leer_precios():
    """
    Lee la hoja Precios del Excel.
    Debe tener columnas: producto, precio
    """
    try:
        df_precios = pd.read_excel(RUTA_ARCHIVO, sheet_name=HOJA_PRECIOS)

        df_precios.columns = [c.strip().lower() for c in df_precios.columns]

        if "producto" not in df_precios.columns or "precio" not in df_precios.columns:
            raise ValueError("La hoja 'Precios' debe contener las columnas 'producto' y 'precio'.")

        df_precios["producto"] = df_precios["producto"].astype(str).str.strip().str.lower()

        precios = dict(zip(df_precios["producto"], df_precios["precio"]))

        faltantes = [p for p in PRODUCTOS_ESPERADOS if p not in precios]
        if faltantes:
            raise ValueError(f"Faltan precios para: {', '.join(faltantes)}")

        return {
            "promocion_completo_bebida": int(precios["promocion_completo_bebida"]),
            "bebida_sola": int(precios["bebida_sola"]),
            "te_solo": int(precios["te_solo"]),
            "cafe_solo": int(precios["cafe_solo"]),
            "completo_solo": int(precios["completo_solo"]),
        }

    except Exception as e:
        raise RuntimeError(f"No fue posible leer la hoja de precios: {e}")


def leer_ventas():
    try:
        df = pd.read_excel(RUTA_ARCHIVO, sheet_name=HOJA_VENTAS)

        for col in COLUMNAS_VENTAS:
            if col not in df.columns:
                df[col] = None

        df = df[COLUMNAS_VENTAS]

        # Eliminar filas completamente vacías
        df = df.dropna(how="all")

        # Eliminar filas donde nombre_comprador esté vacío
        df = df[df["nombre_comprador"].notna()]
        df = df[df["nombre_comprador"].astype(str).str.strip() != ""]

        return df.reset_index(drop=True)

    except Exception as e:
        st.error(f"No fue posible leer las ventas: {e}")
        return pd.DataFrame(columns=COLUMNAS_VENTAS)


def guardar_venta(
    nombre_comprador,
    cantidad_promo,
    cantidad_bebidas,
    cantidad_te,
    cantidad_cafes,
    cantidad_completos,
    precios,
):
    ahora = datetime.now()
    fecha = ahora.strftime("%Y-%m-%d")
    hora = ahora.strftime("%H:%M:%S")

    total_venta = (
        cantidad_promo * precios["promocion_completo_bebida"]
        + cantidad_bebidas * precios["bebida_sola"]
        + cantidad_te * precios["te_solo"]
        + cantidad_cafes * precios["cafe_solo"]
        + cantidad_completos * precios["completo_solo"]
    )

    nueva_fila = [
        fecha,
        hora,
        nombre_comprador.strip(),
        int(cantidad_promo),
        int(cantidad_completos),
        int(cantidad_bebidas),
        int(cantidad_cafes),
        int(cantidad_te),
        int(total_venta),
    ]

    wb = load_workbook(RUTA_ARCHIVO)
    ws = wb[HOJA_VENTAS]

    fila_destino = None

    for fila in range(2, ws.max_row + 2):
        valores = [ws.cell(row=fila, column=col).value for col in range(1, 10)]
        if all(v is None or str(v).strip() == "" for v in valores):
            fila_destino = fila
            break

    if fila_destino is None:
        fila_destino = ws.max_row + 1

    for col_idx, valor in enumerate(nueva_fila, start=1):
        ws.cell(row=fila_destino, column=col_idx, value=valor)

    wb.save(RUTA_ARCHIVO)
    return fila_destino


# ======================================================
# INTERFAZ
# ======================================================
st.title("🧾 Registro de ventas")

if not RUTA_ARCHIVO.exists():
    st.error(f"No se encontró el archivo Excel en la ruta: {RUTA_ARCHIVO}")
    st.stop()

try:
    precios = leer_precios()
except Exception as e:
    st.error(str(e))
    st.stop()

with st.form("formulario_venta", clear_on_submit=True):
    nombre_comprador = st.text_input("Nombre comprador")

    col1, col2 = st.columns(2)

    with col1:
        cantidad_promo = st.number_input(
            "Cantidad promoción completo + bebida",
            min_value=0,
            value=0,
            step=1
        )
        
        
    with col2:

        cantidad_completos = st.number_input(
            "Cantidad de completos solos",
            min_value=0,
            value=0,
            step=1
        )

        cantidad_bebidas = st.number_input(
            "Cantidad de bebidas solas",
            min_value=0,
            value=0,
            step=1
        )

        cantidad_cafes = st.number_input(
            "Cantidad de cafés solos",
            min_value=0,
            value=0,
            step=1
        )

        cantidad_te = st.number_input(
            "Cantidad de té solos",
            min_value=0,
            value=0,
            step=1
        )
        

    total_estimado = (
        int(cantidad_promo) * precios["promocion_completo_bebida"]
        + int(cantidad_bebidas) * precios["bebida_sola"]
        + int(cantidad_te) * precios["te_solo"]
        + int(cantidad_cafes) * precios["cafe_solo"]
        + int(cantidad_completos) * precios["completo_solo"]
    )

    st.markdown(
    f"<h2 style='color:green;'>💰 Total a pagar ({nombre_comprador}): ${total_estimado:,}</h2>".replace(",", "."),
    unsafe_allow_html=True
    )

    guardar = st.form_submit_button("Guardar venta")

    if guardar:
        total_items = (
            int(cantidad_promo)
            + int(cantidad_completos)
            + int(cantidad_bebidas)
            + int(cantidad_te)
            + int(cantidad_cafes)
        )

        if not nombre_comprador.strip():
            st.error("Debes ingresar el nombre del comprador.")
        elif total_items == 0:
            st.error("Debes registrar al menos un producto.")
        else:
            try:
                fila_guardada = guardar_venta(
                    nombre_comprador=nombre_comprador,
                    cantidad_promo=int(cantidad_promo),
                    cantidad_completos=int(cantidad_completos),
                    cantidad_bebidas=int(cantidad_bebidas),
                    cantidad_te=int(cantidad_te),
                    cantidad_cafes=int(cantidad_cafes),
                    precios=precios,
                )
                st.success(f"Venta guardada correctamente en la fila {fila_guardada}.")
            except Exception as e:
                st.error(f"No fue posible guardar la venta: {e}")

st.divider()

st.subheader("Últimas ventas registradas")
df_ventas = leer_ventas()

if df_ventas.empty:
    st.info("Aún no hay ventas registradas.")
else:
    st.dataframe(df_ventas.tail(20), width='stretch', hide_index=True)

with st.expander("💰 Consultar precios vigentes"):
    df_precios_mostrar = pd.DataFrame({
        "Producto": [
            "Promoción completo + bebida",
            "Bebida sola",
            "Té solo",
            "Café solo",
            "Completo solo"
        ],
        "Precio": [
            precios["promocion_completo_bebida"],
            precios["bebida_sola"],
            precios["te_solo"],
            precios["cafe_solo"],
            precios["completo_solo"],
        ]
    })

    st.dataframe(
        df_precios_mostrar,
        width='stretch',
        hide_index=True
    )